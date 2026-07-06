from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from os import replace
from threading import RLock
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet


EXCEL_SCHEMA_VERSION = "2.0"

EXCEL_HEADERS: dict[str, list[str]] = {
    "_meta": ["key", "value", "updated_at"],
    "application_user": ["id", "google_subject_id", "email", "display_name", "profile_image_url", "created_at", "last_login_at"],
    "agent_registry": ["id", "agent_code", "agent_name", "description", "lifecycle_status", "tool_code", "enabled", "created_at", "updated_at"],
    "validation_record": ["id", "external_reference", "source_system", "record_type", "record_title", "expected_agent_code", "payload", "metadata", "active", "created_at"],
    "validation_run": ["id", "record_id", "initiated_by_user_id", "comments", "execution_status", "detected_agent_code", "selected_tool_code", "routing_reason", "routing_confidence", "final_verdict", "final_reason", "final_tag", "final_confidence", "started_at", "completed_at", "error_message"],
    "rule_result": ["id", "run_id", "rule_code", "rule_name", "severity", "passed", "evidence", "message", "tag", "created_at"],
    "llm_judgement": ["id", "run_id", "model_name", "judge_verdict", "confidence", "reason", "findings", "raw_response", "prompt_version", "created_at"],
    "audit_event": ["id", "run_id", "user_id", "event_type", "event_details", "created_at"],
}



def _atomic_save_workbook(workbook: Workbook, target_path: Path) -> None:
    """Save through a temporary file, then replace the target workbook.

    This keeps the Excel-backed POC safer during Streamlit reruns and validation
    writes. It is intentionally simple and local-file only.
    """

    target_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = target_path.with_name(f"{target_path.stem}.tmp{target_path.suffix}")
    workbook.save(temp_path)
    replace(temp_path, target_path)

_excel_lock = RLock()


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def json_dumps(value: Any) -> str:
    return json.dumps(value, default=str, ensure_ascii=False)


def json_loads(value: Any, default: Any | None = None) -> Any:
    if value in (None, ""):
        return {} if default is None else default
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(str(value))
    except json.JSONDecodeError:
        return {} if default is None else default


def initialize_excel_workbook(path: str | Path, reset: bool = False) -> None:
    """Create the Excel workbook and all required sheets if they do not exist."""

    workbook_path = Path(path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    if workbook_path.exists() and not reset:
        wb = load_workbook(workbook_path)
    else:
        wb = Workbook()
        default = wb.active
        wb.remove(default)

    for sheet_name, headers in EXCEL_HEADERS.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx).value = header
                ws.cell(row=1, column=col_idx).font = Font(bold=True)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(len(header) + 4, 16), 36)
        else:
            ws = wb[sheet_name]
            current = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
            if current[: len(headers)] != headers:
                ws.delete_rows(1)
                ws.insert_rows(1)
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=1, column=col_idx).value = header
    meta = wb["_meta"]
    if meta.max_row <= 1:
        meta.append(["schema_version", EXCEL_SCHEMA_VERSION, now_iso()])
        meta.append(["storage_mode", "excel_poc", now_iso()])
        meta.append(["description", "Supervisor Agent Control Tower Excel-backed persistence store", now_iso()])
    _atomic_save_workbook(wb, workbook_path)


class ExcelDataStore:
    """Small Excel persistence layer for the POC.

    It intentionally mirrors the PostgreSQL tables one sheet at a time. The app can
    later move to PostgreSQL by changing STORAGE_BACKEND without touching the UI,
    orchestrator, rules, judge, or synthesizer.
    """

    def __init__(self, path: str | Path):
        self.path = Path(path)
        initialize_excel_workbook(self.path)
        self.workbook = load_workbook(self.path)
        self.dirty = False

    def save(self) -> None:
        if self.dirty:
            self.upsert("_meta", "key", "last_saved_at", {"key": "last_saved_at", "value": now_iso(), "updated_at": now_iso()})
            _atomic_save_workbook(self.workbook, self.path)
            self.dirty = False

    def close(self) -> None:
        self.workbook.close()

    def sheet(self, name: str) -> Worksheet:
        if name not in EXCEL_HEADERS:
            raise ValueError(f"Unknown Excel sheet: {name}")
        return self.workbook[name]

    def headers(self, sheet_name: str) -> list[str]:
        return EXCEL_HEADERS[sheet_name]

    def rows(self, sheet_name: str) -> list[dict[str, Any]]:
        ws = self.sheet(sheet_name)
        headers = self.headers(sheet_name)
        records: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
            if not any(cell not in (None, "") for cell in row):
                continue
            records.append({header: row[idx] for idx, header in enumerate(headers)})
        return records

    def insert(self, sheet_name: str, row: dict[str, Any]) -> None:
        ws = self.sheet(sheet_name)
        headers = self.headers(sheet_name)
        ws.append([self._normalize_cell(row.get(header)) for header in headers])
        self.dirty = True

    def upsert(self, sheet_name: str, key: str, key_value: Any, values: dict[str, Any]) -> dict[str, Any]:
        ws = self.sheet(sheet_name)
        headers = self.headers(sheet_name)
        key_col = headers.index(key) + 1
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=key_col).value == key_value:
                for field, value in values.items():
                    if field in headers:
                        ws.cell(row=row_idx, column=headers.index(field) + 1).value = self._normalize_cell(value)
                self.dirty = True
                return self.find_one(sheet_name, lambda row: row.get(key) == key_value) or values
        self.insert(sheet_name, values)
        return values

    def update(self, sheet_name: str, key: str, key_value: Any, values: dict[str, Any]) -> None:
        ws = self.sheet(sheet_name)
        headers = self.headers(sheet_name)
        key_col = headers.index(key) + 1
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=key_col).value == key_value:
                for field, value in values.items():
                    if field in headers:
                        ws.cell(row=row_idx, column=headers.index(field) + 1).value = self._normalize_cell(value)
                self.dirty = True
                return
        raise ValueError(f"No row found in {sheet_name} where {key}={key_value}")

    def find_one(self, sheet_name: str, predicate: Callable[[dict[str, Any]], bool]) -> dict[str, Any] | None:
        for row in self.rows(sheet_name):
            if predicate(row):
                return row
        return None

    def delete_all(self, sheet_name: str) -> None:
        ws = self.sheet(sheet_name)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
            self.dirty = True

    @staticmethod
    def _normalize_cell(value: Any) -> Any:
        if isinstance(value, (dict, list)):
            return json_dumps(value)
        if isinstance(value, datetime):
            return value.isoformat(timespec="seconds")
        return value


class ExcelTransaction:
    def __init__(self, path: str | Path):
        self.path = Path(path)
        self.store: ExcelDataStore | None = None

    def __enter__(self) -> ExcelDataStore:
        _excel_lock.acquire()
        self.store = ExcelDataStore(self.path)
        return self.store

    def __exit__(self, exc_type: Any, exc: Any, tb: Any) -> None:
        try:
            if self.store is not None and exc_type is None:
                self.store.save()
        finally:
            if self.store is not None:
                self.store.close()
            _excel_lock.release()
